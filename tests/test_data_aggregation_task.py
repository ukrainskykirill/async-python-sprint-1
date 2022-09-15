import openpyxl
from tasks import DataAggregationTask, DataCalculationTask

CITIES = {
    "LONDON": "https://code.s3.yandex.net/async-module/london-response.json",
    "BERLIN": "https://code.s3.yandex.net/async-module/berlin-response.json"
}
CITIES2 = {"LONDON": "https://code.s3.yandex.net/async-module/london-response.json"}
result = [{'city': 'London',
           'data': {'avr_t': {'2022-05-26': 17.36,
                              '2022-05-27': 16.27,
                              '2022-05-28': 14.64,
                              '2022-05-29': None,
                              '2022-05-30': None},
                    'c': 11.0,
                    'condition': {'2022-05-26': 11,
                                  '2022-05-27': 11,
                                  '2022-05-28': 11,
                                  '2022-05-29': 0,
                                  '2022-05-30': None},
                    'raiting': 2,
                    't': 16.09}},
          {'city': 'Teltow-Fläming',
           'data': {'avr_t': {'2022-05-26': 19.27,
                              '2022-05-27': 16.0,
                              '2022-05-28': 13.64,
                              '2022-05-29': None,
                              '2022-05-30': None},
                    'c': 5.0,
                    'condition': {'2022-05-26': 9,
                                  '2022-05-27': 6,
                                  '2022-05-28': 0,
                                  '2022-05-29': 0,
                                  '2022-05-30': None},
                    'raiting': 1,
                    't': 16.3}}]


class TestDataAggregationTask:
    def setup(self):
        self.dc = DataCalculationTask()
        self.da = DataAggregationTask()

    def test_get_from_q(self):
        assert self.da.get_from_q(CITIES) == result

    def test_title(self):
        self.da.title(CITIES2)
        book = openpyxl.open('forecast_test.xlsx', read_only=True)
        sheet = book.active
        first_table = []
        second_table = []
        for row in range(1, sheet.max_row):
            for col in range(0, sheet.max_column):
                first_table.append(sheet[row][col].value)

        book2 = openpyxl.open('forecast.xlsx', read_only=True)
        sheet2 = book2.active
        for row in range(1, sheet2.max_row):
            for col in range(0, sheet2.max_column):
                second_table.append(sheet2[row][col].value)

        assert first_table == second_table
